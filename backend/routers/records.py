from fastapi import APIRouter, Depends, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import Optional
import psycopg
import io

from database import get_db
from auth import get_current_user
from models import RecordCreate, RecordUpdate, ParseRequest
from services.parser import parse_record_text

router = APIRouter(prefix="/api/records", tags=["records"])


@router.get("")
async def list_records(
    search: Optional[str] = Query(None, description="搜索关键词"),
    is_returned: Optional[int] = Query(None, description="回款状态筛选"),
    date_from: Optional[str] = Query(None, description="开始日期 YYYY-MM-DD"),
    date_to: Optional[str] = Query(None, description="结束日期 YYYY-MM-DD"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页条数"),
    db: psycopg.AsyncConnection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """获取记录列表（分页）"""
    user_id = current_user["id"]
    is_admin = current_user["role"] == "admin"

    where = " WHERE 1=1"
    params = []

    # 非管理员只能看自己的记录
    if not is_admin:
        where += " AND user_id = %s"
        params.append(user_id)

    if search:
        where += " AND (customer LIKE %s OR product LIKE %s OR tracking_no LIKE %s OR note LIKE %s)"
        like = f"%{search}%"
        params.extend([like, like, like, like])

    if is_returned is not None:
        where += " AND is_returned = %s"
        params.append(is_returned)

    if date_from:
        where += " AND created_at >= %s"
        params.append(date_from)

    if date_to:
        where += " AND created_at <= %s"
        params.append(date_to + " 23:59:59")

    # 查询总数
    cursor = await db.execute(f"SELECT COUNT(*) as cnt FROM records{where}", params)
    row = await cursor.fetchone()
    total = row["cnt"] if isinstance(row, dict) else row[0]

    # 分页查询
    offset = (page - 1) * page_size
    cursor = await db.execute(
        f"SELECT * FROM records{where} ORDER BY created_at DESC LIMIT %s OFFSET %s",
        params + [page_size, offset],
    )
    rows = await cursor.fetchall()
    return {"data": [dict(row) for row in rows], "total": total, "page": page, "page_size": page_size}


@router.post("")
async def create_record(
    record: RecordCreate,
    db: psycopg.AsyncConnection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """创建新记录"""
    user_id = current_user["id"]
    profit = record.buy_price - record.cost_price + record.other_income

    if record.created_at:
        cursor = await db.execute(
            """INSERT INTO records
               (customer, product, cost_price, buy_price, other_income, profit,
                actual_profit, created_at,
                tracking_no, tracking_company, is_returned, returned_at, note, raw_input, user_id)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
               RETURNING id""",
            (
                record.customer,
                record.product,
                record.cost_price,
                record.buy_price,
                record.other_income,
                profit,
                record.actual_profit,
                record.created_at,
                record.tracking_no,
                record.tracking_company,
                record.is_returned,
                record.returned_at,
                record.note,
                record.raw_input,
                user_id,
            ),
        )
    else:
        cursor = await db.execute(
            """INSERT INTO records
               (customer, product, cost_price, buy_price, other_income, profit,
                actual_profit,
                tracking_no, tracking_company, is_returned, returned_at, note, raw_input, user_id)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
               RETURNING id""",
            (
                record.customer,
                record.product,
                record.cost_price,
                record.buy_price,
                record.other_income,
                profit,
                record.actual_profit,
                record.tracking_no,
                record.tracking_company,
                record.is_returned,
                record.returned_at,
                record.note,
                record.raw_input,
                user_id,
            ),
        )
    row = await cursor.fetchone()
    record_id = row["id"]
    await db.commit()

    cursor = await db.execute("SELECT * FROM records WHERE id = %s", (record_id,))
    row = await cursor.fetchone()
    return {"data": dict(row)}


@router.put("/{record_id}")
async def update_record(
    record_id: int,
    record: RecordUpdate,
    db: psycopg.AsyncConnection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """更新记录"""
    cursor = await db.execute("SELECT * FROM records WHERE id = %s", (record_id,))
    existing = await cursor.fetchone()
    if not existing:
        return {"error": "记录不存在"}

    existing = dict(existing)

    # 权限校验：只能修改自己的记录（admin 除外）
    if current_user["role"] != "admin" and existing.get("user_id") != current_user["id"]:
        return {"error": "无权修改他人的记录"}
    updates = record.model_dump(exclude_unset=True)

    # 如果标记为回款，自动记录回款时间（除非已手动提供）
    auto_returned_at = False
    if "is_returned" in updates and updates["is_returned"] == 1 and not existing["is_returned"]:
        if "returned_at" not in updates or not updates["returned_at"]:
            updates["returned_at"] = "TO_CHAR(NOW(), 'YYYY-MM-DD HH24:MI:SS')"
            auto_returned_at = True

    # 重新计算利润
    cost = updates.get("cost_price", existing["cost_price"])
    buy = updates.get("buy_price", existing["buy_price"])
    other = updates.get("other_income", existing["other_income"])
    updates["profit"] = buy - cost + other
    updates["updated_at"] = "TO_CHAR(NOW(), 'YYYY-MM-DD HH24:MI:SS')"

    set_clauses = []
    values = []
    for k, v in updates.items():
        if k == "updated_at" or (k == "returned_at" and auto_returned_at):
            set_clauses.append(f"{k} = {v}")
        else:
            set_clauses.append(f"{k} = %s")
            values.append(v)

    values.append(record_id)
    await db.execute(
        f"UPDATE records SET {', '.join(set_clauses)} WHERE id = %s",
        values,
    )
    await db.commit()

    cursor = await db.execute("SELECT * FROM records WHERE id = %s", (record_id,))
    row = await cursor.fetchone()
    return {"data": dict(row)}


@router.delete("/{record_id}")
async def delete_record(
    record_id: int,
    db: psycopg.AsyncConnection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """删除记录"""
    # 权限校验：只能删除自己的记录（admin 除外）
    if current_user["role"] != "admin":
        cursor = await db.execute("SELECT user_id FROM records WHERE id = %s", (record_id,))
        row = await cursor.fetchone()
        if not row or row.get("user_id") != current_user["id"]:
            return {"error": "无权删除他人的记录"}

    await db.execute("DELETE FROM records WHERE id = %s", (record_id,))
    await db.commit()
    return {"message": "删除成功"}


@router.post("/parse")
async def parse_text(
    req: ParseRequest,
    current_user: dict = Depends(get_current_user),
):
    """智能解析自然语言输入"""
    # Debug: write input to file
    import os
    debug_path = os.path.join(os.path.dirname(__file__), "..", "_debug_input.txt")
    with open(debug_path, "w", encoding="utf-8") as f:
        f.write(f"INPUT: {repr(req.text)}\n")
        f.write(f"LEN: {len(req.text)}\n")
        f.write(f"CHARS: {[hex(ord(c)) for c in req.text]}\n")
    result = parse_record_text(req.text)
    with open(debug_path, "a", encoding="utf-8") as f:
        f.write(f"PRODUCT: {repr(result.get('product', ''))}\n")
        f.write(f"BUY: {result.get('buy_price', 0)}\n")
    return {"data": result}


@router.get("/import/template")
async def download_import_template(current_user: dict = Depends(get_current_user)):
    """下载导入模板"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "导入模板"
    headers = ["客户", "商品", "成本价", "售价", "其他收入", "实际利润", "快递单号", "快递公司", "是否回款", "下单时间", "备注"]
    ws.append(headers)
    # 示例行
    ws.append(["张三", "RTX5070显卡", 4200, 4500, 0, 300, "SF1234567890", "shunfeng", "否", "2025-06-18", ""])
    # 调整列宽
    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 14
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=import_template.xlsx"},
    )


@router.post("/import")
async def import_records(
    file: UploadFile = File(...),
    db: psycopg.AsyncConnection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """从 Excel 导入记录"""
    from openpyxl import load_workbook

    if not file.filename.endswith((".xlsx", ".xls")):
        return {"error": "请上传 .xlsx 格式的文件"}

    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        return {"error": "表格中没有数据"}

    user_id = current_user["id"]
    imported = 0
    skipped = 0
    errors = []

    for idx, row in enumerate(rows, start=2):
        try:
            # 客户、商品、成本价、售价、其他收入、实际利润、快递单号、快递公司、是否回款、下单时间、备注
            customer = str(row[0] or "").strip()
            product = str(row[1] or "").strip()
            cost_price = float(row[2] or 0)
            buy_price = float(row[3] or 0)
            other_income = float(row[4] or 0)
            actual_profit = float(row[5] or 0)
            tracking_no = str(row[6] or "").strip()
            tracking_company = str(row[7] or "").strip()
            is_returned = 1 if str(row[8] or "").strip() in ("是", "1", "yes", "true") else 0
            created_at = str(row[9] or "").strip() if row[9] else None
            note = str(row[10] or "").strip() if len(row) > 10 else ""

            if not customer and not product:
                skipped += 1
                continue

            profit = buy_price - cost_price + other_income

            await db.execute(
                """INSERT INTO records
                   (customer, product, cost_price, buy_price, other_income, profit,
                    actual_profit, tracking_no, tracking_company, is_returned, note, user_id, created_at)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                (customer, product, cost_price, buy_price, other_income, profit,
                 actual_profit, tracking_no, tracking_company, is_returned, note, user_id,
                 created_at),
            )
            imported += 1
        except Exception as e:
            errors.append(f"第{idx}行: {str(e)}")
            skipped += 1

    await db.commit()
    wb.close()

    return {
        "data": {
            "imported": imported,
            "skipped": skipped,
            "errors": errors[:10],  # 最多返回10条错误
            "message": f"成功导入 {imported} 条记录" + (f"，跳过 {skipped} 条" if skipped else ""),
        }
    }
